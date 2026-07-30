from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .model import ExtractedDocument


@dataclass(slots=True, frozen=True)
class XlsxExportReport:
    pages_exported: int
    tables_exported: int
    fallback_pages: list[int]


def _fit_columns(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        maximum = 0
        for cell in worksheet.iter_cols(
            min_col=column_index, max_col=column_index
        ):
            for item in cell:
                maximum = max(maximum, len(str(item.value or "")))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(maximum + 2, 8), 60
        )


def write_xlsx(
    extracted: ExtractedDocument,
    output_path: str | Path,
) -> XlsxExportReport:
    """Write one worksheet per PDF page, with line fallback when no table exists."""
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    tables_exported = 0
    fallback_pages: list[int] = []

    for page_index, page in enumerate(extracted.pages):
        worksheet = (
            workbook.active
            if page_index == 0
            else workbook.create_sheet()
        )
        worksheet.title = f"Page {page.number}"

        if page.tables:
            output_row = 1
            for table in page.tables:
                table_start_row = output_row
                for row in table.rows:
                    for column_index, value in enumerate(row, start=1):
                        worksheet.cell(output_row, column_index, value)
                    output_row += 1

                for merge in table.merges:
                    worksheet.merge_cells(
                        start_row=table_start_row + merge.min_row - 1,
                        start_column=merge.min_column,
                        end_row=table_start_row + merge.max_row - 1,
                        end_column=merge.max_column,
                    )

                if table.rows:
                    for cell in worksheet[table_start_row]:
                        cell.font = Font(bold=True)
                output_row += 1
                tables_exported += 1
        else:
            fallback_pages.append(page.number)
            for row_index, block in enumerate(page.blocks, start=1):
                worksheet.cell(row_index, 1, block.text)

        worksheet.freeze_panes = "A1"
        _fit_columns(worksheet)

    if not extracted.pages:
        workbook.active.title = "Export"

    workbook.save(target)
    return XlsxExportReport(
        pages_exported=len(extracted.pages),
        tables_exported=tables_exported,
        fallback_pages=fallback_pages,
    )
