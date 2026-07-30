import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from familypdf_office_export.model import (
    ExtractedDocument,
    ExtractedPage,
    ExtractedTable,
    TableMerge,
    TextBlock,
    TextRun,
)
from familypdf_office_export.xlsx_writer import write_xlsx


class XlsxExportTest(unittest.TestCase):
    def test_writes_tables_merges_and_line_fallback(self) -> None:
        document = ExtractedDocument(
            pages=[
                ExtractedPage(
                    number=1,
                    tables=[
                        ExtractedTable(
                            rows=[
                                ["項目", "数量", "金額"],
                                ["家庭測試", 2, 120],
                            ],
                            merges=[TableMerge(1, 1, 1, 2)],
                        ),
                        ExtractedTable(
                            rows=[
                                ["分類", "說明"],
                                ["附加", "第二個表格 / 第二个表格"],
                            ]
                        ),
                    ],
                ),
                ExtractedPage(
                    number=2,
                    blocks=[
                        TextBlock(runs=[TextRun("無表格第一行")]),
                        TextBlock(runs=[TextRun("无表格第二行")]),
                    ],
                ),
            ]
        )

        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "tables.xlsx"
            report = write_xlsx(document, output_path)

            self.assertEqual(report.pages_exported, 2)
            self.assertEqual(report.tables_exported, 2)
            self.assertEqual(report.fallback_pages, [2])
            workbook = load_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["Page 1", "Page 2"])
        first = workbook["Page 1"]
        self.assertEqual(first["A1"].value, "項目")
        self.assertEqual(first["C1"].value, "金額")
        self.assertEqual(first["A2"].value, "家庭測試")
        self.assertEqual(first["B2"].value, 2)
        self.assertEqual(first["C2"].value, 120)
        self.assertIn("A1:B1", {str(item) for item in first.merged_cells.ranges})
        self.assertEqual(first["A4"].value, "分類")
        self.assertEqual(first["B4"].value, "說明")
        self.assertEqual(first["A5"].value, "附加")
        self.assertEqual(first["B5"].value, "第二個表格 / 第二个表格")
        self.assertTrue(first["A1"].font.bold)
        self.assertTrue(first["A4"].font.bold)

        second = workbook["Page 2"]
        self.assertEqual(second["A1"].value, "無表格第一行")
        self.assertEqual(second["A2"].value, "无表格第二行")


if __name__ == "__main__":
    unittest.main()
